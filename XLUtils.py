import openpyxl

def getRowCount(fileName,sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(fileName,sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readData(fileName,sheetName,rowno,columnno):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowno, column=columnno).value

def writeData(fileName,sheetName,rowno,columnno,data):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowno, column=columnno).value = data
    workbook.save(fileName)
